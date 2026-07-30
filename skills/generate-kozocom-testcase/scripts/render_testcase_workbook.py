#!/usr/bin/env python3
"""
Render a Kozocom-format testcase workbook (.xlsx) from a normalized JSON
payload. See references/input-schema.md for the exact JSON contract and
references/template-mapping.md for what every cell means.

Usage:
    python3 render_testcase_workbook.py --input testcases.json --output out.xlsx [--template path.xlsx]
"""
import argparse
import json
import os
from copy import copy

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FIRST_DATA_ROW = 15
HEADER_ANCHOR_ROW = 14
TEMPLATE_LAST_PREFILLED_ROW = 40  # rows 15-40 ship with the No-formula pre-filled
VALID_TYPES = {"Display", "Default", "Function", "Validate", "Transition", "Operation"}
STATUS_COLS = ("J", "O", "T")  # Round 1/2/3 status column letters
ROUND_BLOCK_WIDTH = 5  # Status, BugID, Date, Tester, Memo


def load_payload(path):
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    if "cases" not in payload or not isinstance(payload["cases"], list):
        raise ValueError("Input JSON must have a top-level 'cases' array")
    return payload.get("metadata", {}) or {}, payload["cases"]


def apply_metadata(ws, metadata):
    if metadata.get("environment"):
        ws["G2"] = metadata["environment"]
    if metadata.get("browser"):
        ws["J2"] = metadata["browser"]
    if metadata.get("created_by"):
        ws["B9"] = metadata["created_by"]
    if metadata.get("specs"):
        ws["B10"] = metadata["specs"]
    if metadata.get("url"):
        ws["B12"] = metadata["url"]


def copy_row_style(ws, src_row, dst_row):
    for col_idx in range(1, ws.max_column + 1):
        col = get_column_letter(col_idx)
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        dst.font = copy(src.font)
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format


def write_cases(ws, cases, locale):
    last_row = FIRST_DATA_ROW + len(cases) - 1

    for i, case in enumerate(cases):
        row = FIRST_DATA_ROW + i
        if row > FIRST_DATA_ROW:
            copy_row_style(ws, FIRST_DATA_ROW, row)

        case_type = (case.get("type") or "").strip()
        if case_type not in VALID_TYPES:
            print(f"WARNING: row {row} has invalid type '{case_type}', normalizing to 'Function'. "
                  f"Fix the source JSON to use one of {sorted(VALID_TYPES)}.")
            case_type = "Function"

        ws[f"A{row}"] = locale
        ws[f"B{row}"] = f"=(ROW()-(ROW($B${HEADER_ANCHOR_ROW})))"
        ws[f"C{row}"] = case.get("target_item", "")
        ws[f"D{row}"] = case_type
        ws[f"E{row}"] = case.get("precondition", "")
        ws[f"F{row}"] = case.get("testcase", "")
        ws[f"G{row}"] = case.get("steps", "")
        ws[f"H{row}"] = case.get("data_test", "")
        ws[f"I{row}"] = case.get("expected", "")

        for status_col in STATUS_COLS:
            ws[f"{status_col}{row}"] = "NT"

        for col in ("C", "D", "E", "F", "G", "H", "I"):
            ws[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    return last_row


def clear_tail_rows(ws, last_written_row):
    """Clear the No-formula (and anything else) in rows that ship pre-filled
    in the blank template but were not used, so summary formulas don't count
    them as populated testcases."""
    clear_until = max(TEMPLATE_LAST_PREFILLED_ROW, last_written_row)
    for row in range(last_written_row + 1, clear_until + 1):
        for col_idx in range(1, ws.max_column + 1):
            col = get_column_letter(col_idx)
            ws[f"{col}{row}"] = None


def extend_validations(ws, last_row):
    """Extend Type/Status dropdown validation ranges if cases run past the
    template's pre-validated row 40."""
    if last_row <= TEMPLATE_LAST_PREFILLED_ROW:
        return
    for dv in ws.data_validations.dataValidation:
        for cell_range in dv.sqref.ranges:
            if cell_range.min_row == FIRST_DATA_ROW and cell_range.max_row == TEMPLATE_LAST_PREFILLED_ROW:
                cell_range.max_row = last_row


def render(template_path, input_path, output_path):
    metadata, cases = load_payload(input_path)
    if not cases:
        raise ValueError("'cases' array is empty — nothing to render")

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Template"] if "Template" in wb.sheetnames else wb.worksheets[0]

    apply_metadata(ws, metadata)
    locale = metadata.get("locale", "VN")
    last_row = write_cases(ws, cases, locale)
    clear_tail_rows(ws, last_row)
    extend_validations(ws, last_row)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Rendered {len(cases)} cases to {output_path} (rows {FIRST_DATA_ROW}-{last_row})")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to the normalized testcases JSON")
    parser.add_argument("--output", required=True, help="Path to write the rendered .xlsx")
    parser.add_argument(
        "--template",
        default=None,
        help="Path to the blank Kozocom template (defaults to the bundled asset)",
    )
    args = parser.parse_args()

    template_path = args.template
    if template_path is None:
        here = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(here, "..", "assets", "kozocom-testcase-template.xlsx")

    render(template_path, args.input, args.output)


if __name__ == "__main__":
    main()
