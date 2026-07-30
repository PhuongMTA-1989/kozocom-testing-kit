#!/usr/bin/env python3
"""
Render an API testcase workbook (.xlsx) — Endpoint | No | Type | Precondition |
Scenario | Request Headers | Request Params | Request Body | Expected Status
Code | Expected Response/Behavior — from a normalized JSON payload.

See references/input-schema.md for the JSON contract.

Usage:
    python3 render_api_testcase_workbook.py --input testcases.json --output out.xlsx [--template path.xlsx]
"""
import argparse
import json
import os

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_ROW = 1
FIRST_DATA_ROW = 2
HIGHLIGHT_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
VALID_TYPES = {
    "Positive", "Validation", "Boundary", "Auth", "Permission",
    "Security", "ErrorHandling", "RateLimit", "ContentType", "Regression",
}


def load_items(path):
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    items = payload.get("items") if isinstance(payload, dict) else payload
    if not isinstance(items, list) or not items:
        raise ValueError("Input JSON must contain a non-empty 'items' array (or be a top-level array)")
    return items


def merge_equal_runs(ws, col_idx, first_row, last_row, values):
    col = get_column_letter(col_idx)
    start = first_row
    for row in range(first_row + 1, last_row + 2):
        current = values.get(row)
        previous = values.get(row - 1)
        if row > last_row or current != previous:
            if row - 1 > start:
                ws.merge_cells(f"{col}{start}:{col}{row - 1}")
                ws[f"{col}{start}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            start = row


def render(template_path, input_path, output_path):
    items = load_items(input_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["API_Testcase"] if "API_Testcase" in wb.sheetnames else wb.worksheets[0]

    endpoint_values = {}
    no_counter = 0
    prev_endpoint = None

    for i, item in enumerate(items):
        row = FIRST_DATA_ROW + i
        endpoint = item.get("endpoint", "")
        item_type = (item.get("type") or "").strip()
        if item_type not in VALID_TYPES:
            print(f"WARNING: row {row} has non-standard type '{item_type}'. "
                  f"Expected one of {sorted(VALID_TYPES)} — keeping as-is, but consider normalizing.")

        if endpoint != prev_endpoint:
            no_counter = 0
        no_counter += 1
        prev_endpoint = endpoint

        values = [
            endpoint,
            no_counter,
            item_type,
            item.get("precondition", ""),
            item.get("scenario", ""),
            item.get("headers", ""),
            item.get("params", ""),
            item.get("body", ""),
            item.get("expected_status", ""),
            item.get("expected_response", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="top")

        if item.get("highlight"):
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = HIGHLIGHT_FILL

        endpoint_values[row] = endpoint

    last_row = FIRST_DATA_ROW + len(items) - 1
    merge_equal_runs(ws, 1, FIRST_DATA_ROW, last_row, endpoint_values)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Rendered {len(items)} API testcase rows to {output_path} (rows {FIRST_DATA_ROW}-{last_row})")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to the normalized API testcases JSON")
    parser.add_argument("--output", required=True, help="Path to write the rendered .xlsx")
    parser.add_argument("--template", default=None, help="Path to the blank API testcase template")
    args = parser.parse_args()

    template_path = args.template
    if template_path is None:
        here = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(here, "..", "assets", "api-testcase-template.xlsx")

    render(template_path, args.input, args.output)


if __name__ == "__main__":
    main()
