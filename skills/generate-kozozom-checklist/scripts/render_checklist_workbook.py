#!/usr/bin/env python3
"""
Render a checklist workbook (.xlsx) — Screen/Process | No | Check Point |
Pattern | Expected Result — from a normalized JSON payload.

See references/input-schema.md for the JSON contract.

Usage:
    python3 render_checklist_workbook.py --input checklist.json --output out.xlsx [--template path.xlsx]
"""
import argparse
import json
import os

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_ROW = 1
FIRST_DATA_ROW = 2
HIGHLIGHT_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")


def load_items(path):
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    items = payload.get("items") if isinstance(payload, dict) else payload
    if not isinstance(items, list) or not items:
        raise ValueError("Input JSON must contain a non-empty 'items' array (or be a top-level array)")
    return items


def merge_equal_runs(ws, col_idx, first_row, last_row, values):
    """Merge vertically-consecutive cells in a column that share the same
    value, so repeated Screen/Check Point labels appear once per group like
    the source layout."""
    col = get_column_letter(col_idx)
    start = first_row
    for row in range(first_row + 1, last_row + 2):
        current = values.get(row)
        previous = values.get(row - 1)
        if row > last_row or current != previous:
            if row - 1 > start:
                ws.merge_cells(f"{col}{start}:{col}{row - 1}")
                ws[f"{col}{start}"].alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            start = row


def render(template_path, input_path, output_path):
    items = load_items(input_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Checklist"] if "Checklist" in wb.sheetnames else wb.worksheets[0]

    screen_values = {}
    check_point_values = {}

    no_counter = 0
    prev_screen = None

    for i, item in enumerate(items):
        row = FIRST_DATA_ROW + i
        screen = item.get("screen", "")
        check_point = item.get("check_point", "")
        pattern = item.get("pattern", "")
        expected = item.get("expected_result", "")
        highlight = bool(item.get("highlight", False))

        # Reset the No counter every time we enter a new screen/process block
        if screen != prev_screen:
            no_counter = 0
        no_counter += 1
        prev_screen = screen

        ws.cell(row=row, column=1, value=screen)
        ws.cell(row=row, column=2, value=no_counter)
        ws.cell(row=row, column=3, value=check_point)
        ws.cell(row=row, column=4, value=pattern)
        ws.cell(row=row, column=5, value=expected)

        for col in (2, 3, 4, 5):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="top")

        if highlight:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = HIGHLIGHT_FILL

        screen_values[row] = screen
        check_point_values[row] = check_point

    last_row = FIRST_DATA_ROW + len(items) - 1
    merge_equal_runs(ws, 1, FIRST_DATA_ROW, last_row, screen_values)
    merge_equal_runs(ws, 3, FIRST_DATA_ROW, last_row, check_point_values)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"Rendered {len(items)} checklist rows to {output_path} (rows {FIRST_DATA_ROW}-{last_row})")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to the normalized checklist JSON")
    parser.add_argument("--output", required=True, help="Path to write the rendered .xlsx")
    parser.add_argument(
        "--template",
        default=None,
        help="Path to the blank checklist template (defaults to the bundled asset)",
    )
    args = parser.parse_args()

    template_path = args.template
    if template_path is None:
        here = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(here, "..", "assets", "checklist-template.xlsx")

    render(template_path, args.input, args.output)


if __name__ == "__main__":
    main()
